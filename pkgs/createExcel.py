#createExcel.py
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

# emp table -> excel
wb = Workbook()

sheet1 = wb.active
sheet1.title = '1분기매출'

sheet1.append(['empno', 'ename', 'sal', 'job'])
sheet1.append([7887, 'SMITH', 1000, 'manager'])

# Sheet 2분기 매출 생성
sheet2 = wb.create_sheet("2분기매출")

# 2분기 매출로 옮기기
for row in sheet1.values:
    sheet2.append(row)

wb.save('2019-10-18-myreport.xlsx')+

df = pd.read_excel('2019-10-18-myreport.xlsx', sheet_name='1분기매출', index_col='empno')

print(df)



