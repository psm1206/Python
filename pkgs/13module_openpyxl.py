#13module_openpyxl.py

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

wb = Workbook()
sheet = wb.create_sheet("Mysheet")
sheet.title = "original"

print(wb.sheetnames)
ws = wb.active
# Create cells before accessing a worksheet
# for x in range(1,101):
#     for y in range(1,101):
#         ws.cell(row=x, column=y)

cell_range = sheet['A1':'C2']
# print(cell_range)

# colC = ws['C']
# col_range = ws['C:D']
# row10 = ws[10]
# row_range = ws[5:10]

print(list(sheet.columns))
print(list(sheet.rows))

key = list(sheet.columns)
values = list(sheet.rows)

# ws.values : 해당하는 sheet 'ws'에 있는 모든 값을 row별로 저장.
for row in ws.values:
    for value in row:
        print(value)

# workbook을 저장하겠음.
wb.save('2019-10-18-myreport.xlsx')

wb = load_workbook('titanic.xlsx')

# dataframe
# from openpyxl.utils.dataframe import dataframe_to_rows
# wb = Workbook()
# ws = wb.active
# for r in dataframe_to_rows(df, index=True, header=True):
#     ws.append(r)

# Sheet 값을 Dataframe에 넣는 방법
df = pd.DataFrame(ws.values)

df = pd.DataFrame()
pd.read_excel()
